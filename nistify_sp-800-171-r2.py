#!/usr/bin/env python3
"""
NISTify 800-171 R2 - NIST SP 800-171 Compliance Scanner and Reporter
Windows Compatible Version (No Emojis)
Scans networks and endpoints for compliance with NIST SP 800-171 Rev 2
Generates compliance reports in multiple formats and POA&M documents
"""

import os
import sys
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
import socket
import subprocess
import platform
import datetime
import argparse
import logging
from pathlib import Path
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple
import ipaddress

# Third-party imports
try:
    import nmap
    import pandas as pd
    from jinja2 import Template
    import pdfkit
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import requests
    import networkx as nx
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.patches import FancyBboxPatch
    import numpy as np
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Install with: pip install python-nmap pandas jinja2 pdfkit openpyxl requests networkx matplotlib")
    sys.exit(1)

def print_banner():
    """Print ASCII art banner for NISTify 800-171 R2"""
    banner = """
    ╔══════════════════════════════════════════════════════════════════════════════════════════╗
    ║                                                                                          ║
    ║    ███╗   ██╗██╗███████╗████████╗██╗███████╗██╗   ██╗                                    ║
    ║    ████╗  ██║██║██╔════╝╚══██╔══╝██║██╔════╝╚██╗ ██╔╝                                    ║
    ║    ██╔██╗ ██║██║███████╗   ██║   ██║█████╗   ╚████╔╝                                     ║
    ║    ██║╚██╗██║██║╚════██║   ██║   ██║██╔══╝    ╚██╔╝                                      ║
    ║    ██║ ╚████║██║███████║   ██║   ██║██║        ██║                                       ║
    ║    ╚═╝  ╚═══╝╚═╝╚══════╝   ╚═╝   ╚═╝╚═╝        ╚═╝                                       ║
    ║                                                          By: Nightstalker                ║
    ║              ╔══════════════════════════════════════════════════╗                        ║
    ║              ║             800-171 Rev 2                        ║                        ║
    ║              ╚══════════════════════════════════════════════════╝                        ║
    ║                                                                                          ║
    ║           NIST SP 800-171 Rev 2 Compliance Scanner & Assessment Tool                     ║
    ║                                                                                          ║
    ║   ┌─────────────────────────────────────────────────────────────────────────────────┐    ║
    ║   │  * Automated Network Discovery & Port Scanning                                  │    ║
    ║   │  * NIST SP 800-171 Rev 2 Compliance Assessment                                  │    ║
    ║   │  * SPRS Score Calculation & Risk Analysis                                       │    ║
    ║   │  * Network Topology Visualization                                               │    ║
    ║   │  * Multi-Format Reporting (HTML, JSON, Excel, Text)                             │    ║
    ║   │  * Plan of Action & Milestones (POA&M) Generation                               │    ║
    ║   └─────────────────────────────────────────────────────────────────────────────────┘    ║
    ║                                                                                          ║
    ║               Version: 1.2.0  |  License: MIT  |  Windows Compatible                     ║
    ║                                                                                          ║
    ╚══════════════════════════════════════════════════════════════════════════════════════════╝

    """
    print(banner)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('nistify800-171r2.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class ComplianceResult:
    """Result of a compliance check"""
    control_id: str
    control_name: str
    control_text: str
    status: str
    finding: str
    remediation: str
    severity: str
    evidence: List[str]

@dataclass
class NetworkTopology:
    """Network topology information"""
    nodes: Dict[str, Dict]
    edges: List[Tuple[str, str]]
    subnets: List[str]
    gateways: List[str]
    network_diagram_path: Optional[str] = None

@dataclass
class SystemInfo:
    """System information for scanned endpoint"""
    hostname: str
    ip_address: str
    os_type: str
    os_version: str
    open_ports: List[int]
    services: Dict[int, str]
    last_scanned: str
    mac_address: Optional[str] = None
    vendor: Optional[str] = None
    hop_count: Optional[int] = None
    gateway: Optional[str] = None

class NIST80171Controls:
    """NIST SP 800-171 Rev 2 Control Definitions"""
    
    CONTROLS = {
        "3.1.1": {
        "name": "Access Control Policy and Procedures",
        "text": "Limit system access to authorized users, processes acting on behalf of authorized users, and devices (including other systems).",
        "family": "AC"
    },
    "3.1.2": {
        "name": "Transaction and Function Control",
        "text": "Limit system access to the types of transactions and functions that authorized users are permitted to execute.",
        "family": "AC"
    },
    "3.1.3": {
        "name": "External Connections",
        "text": "Control the flow of CUI in accordance with approved authorizations.",
        "family": "AC"
    },
    "3.1.4": {
        "name": "Separation of Duties",
        "text": "Separate the duties of individuals to reduce the risk of malevolent activity without collusion.",
        "family": "AC"
    },
    "3.1.5": {
        "name": "Least Privilege",
        "text": "Employ the principle of least privilege, including for specific security functions and privileged accounts.",
        "family": "AC"
    },
    "3.1.6": {
        "name": "Non-Privileged Account Use",
        "text": "Use non-privileged accounts or roles when accessing nonsecurity functions.",
        "family": "AC"
    },
    "3.1.7": {
        "name": "Privileged Functions",
        "text": "Prevent non-privileged users from executing privileged functions and capture the execution of such functions in audit logs.",
        "family": "AC"
    },
    "3.1.8": {
        "name": "Unsuccessful Logon Attempts",
        "text": "Limit unsuccessful logon attempts.",
        "family": "AC"
    },
    "3.1.9": {
        "name": "Privacy and Security Notices",
        "text": "Provide privacy and security notices consistent with applicable CUI rules.",
        "family": "AC"
    },
    "3.1.10": {
        "name": "Session Lock",
        "text": "Use session lock with pattern-hiding displays to prevent access and viewing of data after a period of inactivity.",
        "family": "AC"
    },
    "3.1.11": {
        "name": "Session Termination",
        "text": "Terminate (automatically) a user session after a defined condition.",
        "family": "AC"
    },
    "3.1.12": {
        "name": "Control Remote Access",
        "text": "Monitor and control remote access sessions.",
        "family": "AC"
    },
    "3.1.13": {
        "name": "Remote Access Confidentiality",
        "text": "Employ cryptographic mechanisms to protect the confidentiality of remote access sessions.",
        "family": "AC"
    },
    "3.1.14": {
        "name": "Remote Access Routing",
        "text": "Route remote access via managed access control points.",
        "family": "AC"
    },
    "3.1.15": {
        "name": "Privileged Remote Access",
        "text": "Authorize remote execution of privileged commands and remote access to security-relevant information.",
        "family": "AC"
    },
    "3.1.16": {
        "name": "Wireless Access Authorization",
        "text": "Authorize wireless access prior to allowing such connections.",
        "family": "AC"
    },
    "3.1.17": {
        "name": "Wireless Access Protection",
        "text": "Protect wireless access using authentication and encryption.",
        "family": "AC"
    },
    "3.1.18": {
        "name": "Mobile Device Connection",
        "text": "Control connection of mobile devices.",
        "family": "AC"
    },
    "3.1.19": {
        "name": "Encrypt CUI on Mobile",
        "text": "Encrypt CUI on mobile devices and mobile computing platforms.",
        "family": "AC"
    },
    "3.1.20": {
        "name": "External System Use",
        "text": "Verify and control/limit connections to and use of external systems.",
        "family": "AC"
    },
    "3.1.21": {
        "name": "Portable Storage Device Use",
        "text": "Limit use of portable storage devices on external systems.",
        "family": "AC"
    },
    "3.1.22": {
        "name": "Publicly Accessible System Content",
        "text": "Control CUI posted or processed on publicly accessible systems.",
        "family": "AC"
    },
    
    # 3.2 AWARENESS AND TRAINING (AT)
    "3.2.1": {
        "name": "Security Awareness",
        "text": "Ensure that managers, systems administrators, and users of organizational systems are made aware of the security risks associated with their activities and of the applicable policies, standards, and procedures related to the security of those systems.",
        "family": "AT"
    },
    "3.2.2": {
        "name": "Insider Threat Awareness",
        "text": "Ensure that personnel are trained to carry out their assigned information security-related duties and responsibilities.",
        "family": "AT"
    },
    "3.2.3": {
        "name": "Security Training",
        "text": "Provide security awareness training on recognizing and reporting potential indicators of insider threat.",
        "family": "AT"
    },
    
    # 3.3 AUDIT AND ACCOUNTABILITY (AU)
    "3.3.1": {
        "name": "System Audit Records",
        "text": "Create and retain system audit logs and records to the extent needed to enable the monitoring, analysis, investigation, and reporting of unlawful or unauthorized system activity.",
        "family": "AU"
    },
    "3.3.2": {
        "name": "Audit Record Content",
        "text": "Ensure that the actions of individual system users can be uniquely traced to those users so they can be held accountable for their actions.",
        "family": "AU"
    },
    "3.3.3": {
        "name": "Audit Record Review",
        "text": "Review and update logged events.",
        "family": "AU"
    },
    "3.3.4": {
        "name": "Audit Failure Response",
        "text": "Alert in the event of an audit logging process failure.",
        "family": "AU"
    },
    "3.3.5": {
        "name": "Audit Correlation",
        "text": "Correlate audit record review, analysis, and reporting processes for investigation and response to indications of unlawful, unauthorized, suspicious, or unusual activity.",
        "family": "AU"
    },
    "3.3.6": {
        "name": "Audit Reduction",
        "text": "Provide audit record reduction and report generation to support on-demand analysis and reporting.",
        "family": "AU"
    },
    "3.3.7": {
        "name": "Audit Monitoring",
        "text": "Provide a system capability that compares and synchronizes internal system clocks with an authoritative source to generate time stamps for audit records.",
        "family": "AU"
    },
    "3.3.8": {
        "name": "Audit Record Protection",
        "text": "Protect audit information and audit logging tools from unauthorized access, modification, and deletion.",
        "family": "AU"
    },
    "3.3.9": {
        "name": "Audit Management",
        "text": "Limit management of audit logging functionality to a subset of privileged users.",
        "family": "AU"
    },
    
    # 3.4 CONFIGURATION MANAGEMENT (CM)
    "3.4.1": {
        "name": "Baseline Configuration",
        "text": "Establish and maintain baseline configurations and inventories of organizational systems (including hardware, software, firmware, and documentation) throughout the respective system development life cycles.",
        "family": "CM"
    },
    "3.4.2": {
        "name": "Security Configuration Settings",
        "text": "Establish and enforce security configuration settings for information technology products employed in organizational systems.",
        "family": "CM"
    },
    "3.4.3": {
        "name": "Configuration Change Control",
        "text": "Track, review, approve or disapprove, and log changes to organizational systems.",
        "family": "CM"
    },
    "3.4.4": {
        "name": "Security Impact Analysis",
        "text": "Analyze the security impact of changes prior to implementation.",
        "family": "CM"
    },
    "3.4.5": {
        "name": "Access Restrictions",
        "text": "Define, document, approve, and enforce physical and logical access restrictions associated with changes to organizational systems.",
        "family": "CM"
    },
    "3.4.6": {
        "name": "Least Functionality",
        "text": "Employ the principle of least functionality by configuring organizational systems to provide only essential capabilities.",
        "family": "CM"
    },
    "3.4.7": {
        "name": "Nonessential Functionality",
        "text": "Restrict, disable, or prevent the use of nonessential programs, functions, ports, protocols, and services.",
        "family": "CM"
    },
    "3.4.8": {
        "name": "Application Execution Policy",
        "text": "Apply deny-by-exception (blacklisting) policy to prevent the use of unauthorized software or deny-all, permit-by-exception (whitelisting) policy to allow the execution of authorized software.",
        "family": "CM"
    },
    "3.4.9": {
        "name": "User-Installed Software",
        "text": "Control and monitor user-installed software.",
        "family": "CM"
    },
    
    # 3.5 IDENTIFICATION AND AUTHENTICATION (IA)
    "3.5.1": {
        "name": "User Identification",
        "text": "Identify system users, processes acting on behalf of users, and devices.",
        "family": "IA"
    },
    "3.5.2": {
        "name": "User Authentication",
        "text": "Authenticate (or verify) the identities of users, processes, or devices, as a prerequisite to allowing access to organizational systems.",
        "family": "IA"
    },
    "3.5.3": {
        "name": "Multifactor Authentication",
        "text": "Use multifactor authentication for local and network access to privileged accounts and for network access to non-privileged accounts.",
        "family": "IA"
    },
    "3.5.4": {
        "name": "Replay-Resistant Authentication",
        "text": "Employ replay-resistant authentication mechanisms for network access to privileged and non-privileged accounts.",
        "family": "IA"
    },
    "3.5.5": {
        "name": "Identifier Management",
        "text": "Prevent reuse of identifiers for a defined period.",
        "family": "IA"
    },
    "3.5.6": {
        "name": "Authenticator Management",
        "text": "Disable identifiers after a defined period of inactivity.",
        "family": "IA"
    },
    "3.5.7": {
        "name": "Password Complexity",
        "text": "Enforce a minimum password complexity and change of characters when new passwords are created.",
        "family": "IA"
    },
    "3.5.8": {
        "name": "Password Reuse",
        "text": "Prohibit password reuse for a specified number of generations.",
        "family": "IA"
    },
    "3.5.9": {
        "name": "Temporary Password",
        "text": "Allow temporary password use for system logons with an immediate change to a permanent password.",
        "family": "IA"
    },
    "3.5.10": {
        "name": "Cryptographic Authentication",
        "text": "Store and transmit only cryptographically-protected passwords.",
        "family": "IA"
    },
    "3.5.11": {
        "name": "Obscure Feedback",
        "text": "Obscure feedback of authentication information.",
        "family": "IA"
    },
    
    # 3.6 INCIDENT RESPONSE (IR)
    "3.6.1": {
        "name": "Incident Response Plan",
        "text": "Establish an operational incident-handling capability for organizational systems that includes preparation, detection, analysis, containment, recovery, and user response activities.",
        "family": "IR"
    },
    "3.6.2": {
        "name": "Incident Tracking",
        "text": "Track, document, and report incidents to designated officials and/or authorities both internal and external to the organization.",
        "family": "IR"
    },
    "3.6.3": {
        "name": "Incident Testing",
        "text": "Test the organizational incident response capability.",
        "family": "IR"
    },
    
    # 3.7 MAINTENANCE (MA)
    "3.7.1": {
        "name": "Maintenance Policy",
        "text": "Perform maintenance on organizational systems.",
        "family": "MA"
    },
    "3.7.2": {
        "name": "Controlled Maintenance",
        "text": "Provide controls on the tools, techniques, mechanisms, and personnel used to conduct system maintenance.",
        "family": "MA"
    },
    "3.7.3": {
        "name": "Maintenance Tools",
        "text": "Ensure equipment removed for off-site maintenance is sanitized of any CUI.",
        "family": "MA"
    },
    "3.7.4": {
        "name": "Nonlocal Maintenance",
        "text": "Check media containing diagnostic and test programs for malicious code before the media are used in organizational systems.",
        "family": "MA"
    },
    "3.7.5": {
        "name": "Maintenance Personnel",
        "text": "Require multifactor authentication to establish nonlocal maintenance sessions via external network connections and terminate such connections when nonlocal maintenance is complete.",
        "family": "MA"
    },
    "3.7.6": {
        "name": "Maintenance Supervision",
        "text": "Supervise the maintenance activities of maintenance personnel without required access authorization.",
        "family": "MA"
    },
    
    # 3.8 MEDIA PROTECTION (MP)
    "3.8.1": {
        "name": "Media Access",
        "text": "Protect (i.e., physically control and securely store) system media containing CUI, both paper and digital.",
        "family": "MP"
    },
    "3.8.2": {
        "name": "Media Disposal",
        "text": "Limit access to CUI on system media to authorized users.",
        "family": "MP"
    },
    "3.8.3": {
        "name": "Media Sanitization",
        "text": "Sanitize or destroy system media containing CUI before disposal or release for reuse.",
        "family": "MP"
    },
    "3.8.4": {
        "name": "Media Marking",
        "text": "Mark media with necessary CUI markings and distribution limitations.",
        "family": "MP"
    },
    "3.8.5": {
        "name": "Media Transport",
        "text": "Control access to media containing CUI and maintain accountability for media during transport outside of controlled areas.",
        "family": "MP"
    },
    "3.8.6": {
        "name": "Cryptographic Protection",
        "text": "Implement cryptographic mechanisms to protect the confidentiality of CUI stored on digital media during transport unless otherwise protected by alternative physical safeguards.",
        "family": "MP"
    },
    "3.8.7": {
        "name": "Portable Storage",
        "text": "Control the use of removable media on system components.",
        "family": "MP"
    },
    "3.8.8": {
        "name": "Media Downgrading",
        "text": "Prohibit the use of portable storage devices when such devices have no identifiable owner.",
        "family": "MP"
    },
    "3.8.9": {
        "name": "Media Protection",
        "text": "Protect the confidentiality of backup CUI at storage locations.",
        "family": "MP"
    },
    
    # 3.9 PERSONNEL SECURITY (PS)
    "3.9.1": {
        "name": "Personnel Screening",
        "text": "Screen individuals prior to authorizing access to organizational systems containing CUI.",
        "family": "PS"
    },
    "3.9.2": {
        "name": "Personnel Termination",
        "text": "Ensure that organizational systems containing CUI are protected during and after personnel actions such as terminations and transfers.",
        "family": "PS"
    },
    
    # 3.10 PHYSICAL PROTECTION (PE)
    "3.10.1": {
        "name": "Physical Access Authorizations",
        "text": "Limit physical access to organizational systems, equipment, and the respective operating environments to authorized individuals.",
        "family": "PE"
    },
    "3.10.2": {
        "name": "Physical Access Controls",
        "text": "Protect and monitor the physical facility and support infrastructure for organizational systems.",
        "family": "PE"
    },
    "3.10.3": {
        "name": "Escort Visitors",
        "text": "Escort visitors and monitor visitor activity.",
        "family": "PE"
    },
    "3.10.4": {
        "name": "Physical Access Logs",
        "text": "Maintain audit logs of physical access.",
        "family": "PE"
    },
    "3.10.5": {
        "name": "Physical Access Devices",
        "text": "Control and manage physical access devices.",
        "family": "PE"
    },
    "3.10.6": {
        "name": "Monitoring Physical Access",
        "text": "Enforce safeguarding measures for CUI at alternate work sites.",
        "family": "PE"
    },
    
    # 3.11 RISK ASSESSMENT (RA)
    "3.11.1": {
        "name": "Risk Assessment",
        "text": "Periodically assess the risk to organizational operations (including mission, functions, image, or reputation), organizational assets, and individuals, resulting from the operation of organizational systems and the associated processing, storage, or transmission of CUI.",
        "family": "RA"
    },
    "3.11.2": {
        "name": "Vulnerability Scanning",
        "text": "Scan for vulnerabilities in organizational systems and applications periodically and when new vulnerabilities affecting those systems and applications are identified.",
        "family": "RA"
    },
    "3.11.3": {
        "name": "Remediation",
        "text": "Remediate vulnerabilities in accordance with risk assessments.",
        "family": "RA"
    },
    
    # 3.12 SECURITY ASSESSMENT (CA)
    "3.12.1": {
        "name": "Security Assessments",
        "text": "Periodically assess the security controls in organizational systems to determine if the controls are effective in their application.",
        "family": "CA"
    },
    "3.12.2": {
        "name": "Plans of Action",
        "text": "Develop and implement plans of action designed to correct deficiencies and reduce or eliminate vulnerabilities in organizational systems.",
        "family": "CA"
    },
    "3.12.3": {
        "name": "System Interconnections",
        "text": "Monitor security controls on an ongoing basis to ensure the continued effectiveness of the controls.",
        "family": "CA"
    },
    "3.12.4": {
        "name": "Security Control Testing",
        "text": "Develop, document, and periodically update system security plans that describe system boundaries, system environments of operation, how security requirements are implemented, and the relationships with or connections to other systems.",
        "family": "CA"
    },
    
    # 3.13 SYSTEM AND COMMUNICATIONS PROTECTION (SC)
    "3.13.1": {
        "name": "Boundary Protection",
        "text": "Monitor, control, and protect communications (i.e., information transmitted or received by organizational systems) at the external boundaries and key internal boundaries of organizational systems.",
        "family": "SC"
    },
    "3.13.2": {
        "name": "Application Partitioning",
        "text": "Employ architectural designs, software development techniques, and systems engineering principles that promote effective information security within organizational systems.",
        "family": "SC"
    },
    "3.13.3": {
        "name": "Security Function Isolation",
        "text": "Separate user functionality from system management functionality.",
        "family": "SC"
    },
    "3.13.4": {
        "name": "Information in Shared Resources",
        "text": "Prevent unauthorized and unintended information transfer via shared system resources.",
        "family": "SC"
    },
    "3.13.5": {
        "name": "Denial of Service Protection",
        "text": "Implement subnetworks for publicly accessible system components that are physically or logically separated from internal networks.",
        "family": "SC"
    },
    "3.13.6": {
        "name": "Network Segmentation",
        "text": "Deny network communications traffic by default and allow network communications traffic by exception (i.e., deny all, permit by exception).",
        "family": "SC"
    },
    "3.13.7": {
        "name": "Split Tunneling",
        "text": "Prevent remote devices from simultaneously establishing non-remote connections with organizational systems and communicating via some other connection to resources in external networks (i.e., split tunneling).",
        "family": "SC"
    },
    "3.13.8": {
        "name": "Cryptographic Protection",
        "text": "Implement cryptographic mechanisms to prevent unauthorized disclosure of CUI during transmission unless otherwise protected by alternative physical safeguards.",
        "family": "SC"
    },
    "3.13.9": {
        "name": "Session Termination",
        "text": "Terminate network connections associated with communications sessions at the end of the sessions or after a defined period of inactivity.",
        "family": "SC"
    },
    "3.13.10": {
        "name": "Cryptographic Key Management",
        "text": "Establish and manage cryptographic keys for cryptography employed in organizational systems.",
        "family": "SC"
    },
    "3.13.11": {
        "name": "CUI Confidentiality",
        "text": "Employ FIPS-validated cryptography when used to protect the confidentiality of CUI.",
        "family": "SC"
    },
    "3.13.12": {
        "name": "Collaborative Computing Devices",
        "text": "Prohibit remote activation of collaborative computing devices and provide indication of devices in use to users present at the device.",
        "family": "SC"
    },
    "3.13.13": {
        "name": "Mobile Code",
        "text": "Control and monitor the use of mobile code.",
        "family": "SC"
    },
    "3.13.14": {
        "name": "Voice over Internet Protocol",
        "text": "Control and monitor the use of Voice over Internet Protocol (VoIP) technologies.",
        "family": "SC"
    },
    "3.13.15": {
        "name": "Authenticity Protection",
        "text": "Protect the authenticity of communications sessions.",
        "family": "SC"
    },
    "3.13.16": {
        "name": "Transmission Confidentiality",
        "text": "Protect the confidentiality of CUI at rest.",
        "family": "SC"
    },
    
    # 3.14 SYSTEM AND INFORMATION INTEGRITY (SI)
    "3.14.1": {
        "name": "Flaw Remediation",
        "text": "Identify, report, and correct system flaws in a timely manner.",
        "family": "SI"
    },
    "3.14.2": {
        "name": "Malicious Code Protection",
        "text": "Provide protection from malicious code at designated locations within organizational systems.",
        "family": "SI"
    },
    "3.14.3": {
        "name": "Security Alerts and Advisories",
        "text": "Monitor system security alerts and advisories and take action in response.",
        "family": "SI"
    },
    "3.14.4": {
        "name": "Software and Firmware Integrity",
        "text": "Update malicious code protection mechanisms when new releases are available.",
        "family": "SI"
    },
    "3.14.5": {
        "name": "Spam Protection",
        "text": "Perform periodic scans of organizational systems and real-time scans of files from external sources as files are downloaded, opened, or executed.",
        "family": "SI"
    },
    "3.14.6": {
        "name": "Information Handling and Retention",
        "text": "Monitor organizational systems, including inbound and outbound communications traffic, to detect attacks and indicators of potential attacks.",
        "family": "SI"
    },
    "3.14.7": {
        "name": "Information System Monitoring",
        "text": "Identify unauthorized use of organizational systems.",
        "family": "SI"
        }
    }

class NetworkScanner:
    """Network and endpoint scanner using nmap"""
    
    def __init__(self):
        try:
            self.nm = nmap.PortScanner()
            self.topology = None
        except nmap.PortScannerError as e:
            logger.error(f"Nmap initialization failed: {e}")
            logger.error("Please install nmap:")
            logger.error("Windows: Download from https://nmap.org/download.html")
            logger.error("Linux: sudo apt-get install nmap")
            logger.error("macOS: brew install nmap")
            raise SystemExit("Nmap is required but not found. Please install nmap and try again.")
        
    def scan_network(self, network_range: str, ports: str = "22,23,53,80,135,139,443,445,993,995") -> List[SystemInfo]:
        """Scan network range for active hosts and services"""
        logger.info(f"[SCAN] Scanning network range: {network_range}")
        systems = []
        
        try:
            is_windows = platform.system().lower() == 'windows'
            
            if is_windows:
                scan_args = '-sn'
            else:
                scan_args = '-sn -PR -PS21,22,23,25,53,80,110,111,135,139,143,443,993,995'
            
            self.nm.scan(hosts=network_range, arguments=scan_args)
            active_hosts = [host for host in self.nm.all_hosts() if self.nm[host].state() == 'up']
            
            logger.info(f"[SCAN] Found {len(active_hosts)} active hosts")
            
            for host in active_hosts:
                try:
                    if is_windows:
                        self.nm.scan(host, ports, arguments='-sV -sS')
                    else:
                        self.nm.scan(host, ports, arguments='-sV -O -A --version-all')
                    
                    if host in self.nm.all_hosts():
                        host_info = self.nm[host]
                        
                        os_type = "Unknown"
                        os_version = "Unknown"
                        
                        if 'osmatch' in host_info and host_info['osmatch']:
                            os_match = host_info['osmatch'][0]
                            os_type = os_match.get('name', 'Unknown')
                            if 'osclass' in os_match and os_match['osclass']:
                                os_version = os_match['osclass'][0].get('osfamily', 'Unknown')
                        
                        open_ports = []
                        services = {}
                        
                        if 'tcp' in host_info:
                            for port, port_info in host_info['tcp'].items():
                                if port_info['state'] == 'open':
                                    open_ports.append(port)
                                    service_name = port_info.get('name', 'unknown')
                                    service_version = port_info.get('version', '')
                                    services[port] = f"{service_name} {service_version}".strip()
                        
                        hostname = host_info.hostname() if host_info.hostname() else host
                        mac_address = None
                        vendor = None
                        
                        if 'addresses' in host_info:
                            addresses = host_info['addresses']
                            if 'mac' in addresses:
                                mac_address = addresses['mac']
                        
                        if 'vendor' in host_info and host_info['vendor']:
                            vendor = list(host_info['vendor'].values())[0] if host_info['vendor'] else None
                        
                        hop_count = None
                        gateway = None
                        if not is_windows and 'traceroute' in host_info:
                            traceroute = host_info['traceroute']
                            if traceroute:
                                hop_count = len(traceroute)
                                if len(traceroute) > 0:
                                    gateway = traceroute[0].get('ipaddr')
                        else:
                            gateway = self._get_default_gateway_windows() if is_windows else None
                        
                        system_info = SystemInfo(
                            hostname=hostname,
                            ip_address=host,
                            os_type=os_type,
                            os_version=os_version,
                            open_ports=open_ports,
                            services=services,
                            last_scanned=datetime.datetime.now().isoformat(),
                            mac_address=mac_address,
                            vendor=vendor,
                            hop_count=hop_count,
                            gateway=gateway
                        )
                        
                        systems.append(system_info)
                        logger.info(f"[SCAN] Scanned {host}: {len(open_ports)} open ports, OS: {os_type}")
                        
                except Exception as e:
                    logger.error(f"[ERROR] Error scanning host {host}: {e}")
                    continue
                    
        except Exception as e:
            logger.error(f"[ERROR] Error during network scan: {e}")
            
        return systems
    
    def _get_default_gateway_windows(self) -> Optional[str]:
        """Get default gateway on Windows systems"""
        try:
            result = subprocess.run(['ipconfig'], capture_output=True, text=True, shell=True)
            lines = result.stdout.split('\n')
            for line in lines:
                if 'Default Gateway' in line and ':' in line:
                    gateway = line.split(':')[1].strip()
                    if gateway and gateway != '':
                        return gateway
        except Exception as e:
            logger.debug(f"Could not determine default gateway: {e}")
        return None
    
    def discover_topology(self, systems: List[SystemInfo]) -> NetworkTopology:
        """Discover and map network topology from scan results"""
        logger.info("[TOPOLOGY] Analyzing network topology...")
        
        nodes = {}
        edges = []
        subnets = set()
        gateways = set()
        
        for system in systems:
            node_info = {
                'hostname': system.hostname,
                'ip': system.ip_address,
                'os_type': system.os_type,
                'open_ports': len(system.open_ports),
                'services': list(system.services.values())[:3],
                'mac_address': system.mac_address,
                'vendor': system.vendor,
                'hop_count': system.hop_count or 1,
                'type': self._classify_node_type(system)
            }
            nodes[system.ip_address] = node_info
            
            try:
                network = ipaddress.ip_network(f"{system.ip_address}/24", strict=False)
                subnets.add(str(network))
            except:
                pass
            
            if system.gateway:
                gateways.add(system.gateway)
                if system.gateway != system.ip_address:
                    edges.append((system.ip_address, system.gateway))
        
        for gateway in gateways:
            if gateway not in nodes:
                nodes[gateway] = {
                    'hostname': f'Gateway-{gateway}',
                    'ip': gateway,
                    'os_type': 'Gateway/Router',
                    'open_ports': 0,
                    'services': ['Routing'],
                    'mac_address': None,
                    'vendor': 'Unknown',
                    'hop_count': 0,
                    'type': 'gateway'
                }
        
        topology = NetworkTopology(
            nodes=nodes,
            edges=edges,
            subnets=list(subnets),
            gateways=list(gateways)
        )
        
        self.topology = topology
        return topology
    
    def _classify_node_type(self, system: SystemInfo) -> str:
        """Classify node type based on services and characteristics"""
        services = [service.lower() for service in system.services.values()]
        open_ports = system.open_ports
        
        if any('http' in service or 'web' in service for service in services):
            return 'web_server'
        elif 22 in open_ports or any('ssh' in service for service in services):
            return 'server'
        elif 'windows' in system.os_type.lower():
            return 'windows_client'
        elif 'linux' in system.os_type.lower():
            return 'linux_client'
        
        return 'unknown'
    
    def create_network_diagram(self, topology: NetworkTopology, output_path: str = "network_topology.png"):
        """Create a visual network topology diagram"""
        logger.info("[DIAGRAM] Generating network topology diagram...")
        
        try:
            G = nx.Graph()
            
            for ip, node_info in topology.nodes.items():
                G.add_node(ip, **node_info)
            
            G.add_edges_from(topology.edges)
            
            plt.figure(figsize=(16, 12))
            plt.clf()
            
            node_colors = {
                'gateway': '#FF6B6B',
                'web_server': '#4ECDC4',
                'server': '#45B7D1',
                'windows_client': '#AED6F1',
                'linux_client': '#A9DFBF',
                'unknown': '#D5DBDB'
            }
            
            pos = nx.spring_layout(G, k=3, iterations=50, seed=42)
            
            for node_type, color in node_colors.items():
                nodes_of_type = [node for node, data in G.nodes(data=True) if data.get('type') == node_type]
                if nodes_of_type:
                    node_sizes = [1000 + (G.nodes[node].get('open_ports', 0) * 100) for node in nodes_of_type]
                    nx.draw_networkx_nodes(G, pos, nodelist=nodes_of_type, 
                                         node_color=color, node_size=node_sizes, 
                                         alpha=0.8, edgecolors='black', linewidths=1)
            
            nx.draw_networkx_edges(G, pos, alpha=0.6)
            
            labels = {}
            for node, data in G.nodes(data=True):
                hostname = data.get('hostname', node)
                if hostname != node:
                    labels[node] = f"{hostname}\n{node}"
                else:
                    labels[node] = node
            
            nx.draw_networkx_labels(G, pos, labels, font_size=8, font_weight='bold')
            
            legend_elements = []
            for node_type, color in node_colors.items():
                if any(data.get('type') == node_type for _, data in G.nodes(data=True)):
                    legend_elements.append(mpatches.Patch(color=color, label=node_type.replace('_', ' ').title()))
            
            plt.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(0, 1))
            
            plt.title("Network Topology Diagram\nNISTify 800-171 R2 Compliance Assessment", 
                     fontsize=16, fontweight='bold', pad=20)
            
            info_text = f"Total Nodes: {len(G.nodes())}\nTotal Connections: {len(G.edges())}\n"
            info_text += f"Subnets: {len(topology.subnets)}\nGateways: {len(topology.gateways)}"
            
            plt.text(0.02, 0.98, info_text, transform=plt.gca().transAxes, 
                    verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
            
            plt.axis('off')
            plt.tight_layout()
            plt.savefig(output_path, dpi=300, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            plt.close()
            
            topology.network_diagram_path = output_path
            logger.info(f"[DIAGRAM] Network topology diagram saved: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"[ERROR] Error creating network diagram: {e}")
            return None

class ComplianceAssessor:
    """Assess NIST SP 800-171 compliance based on scan results"""
    
    def __init__(self):
        self.controls = NIST80171Controls.CONTROLS
        
    def assess_system(self, system: SystemInfo) -> List[ComplianceResult]:
        """Assess a single system for compliance"""
        results = []
        
        logger.info(f"[ASSESS] Assessing compliance for {system.hostname} ({system.ip_address})")
        
        weak_services = self._check_weak_services(system)
        if weak_services:
            results.append(ComplianceResult(
                control_id="3.1.2",
                control_name=self.controls["3.1.2"]["name"],
                control_text=self.controls["3.1.2"]["text"],
                status="non_compliant",
                finding=f"Potentially insecure services detected: {', '.join(weak_services)}",
                remediation="Disable unnecessary services, implement strong authentication, and restrict access",
                severity="high",
                evidence=[f"Open ports: {system.open_ports}", f"Services: {system.services}"]
            ))
        else:
            results.append(ComplianceResult(
                control_id="3.1.2",
                control_name=self.controls["3.1.2"]["name"],
                control_text=self.controls["3.1.2"]["text"],
                status="compliant",
                finding="No obviously insecure services detected",
                remediation="Continue monitoring for unauthorized services",
                severity="low",
                evidence=[f"Services reviewed: {list(system.services.values())}"]
            ))
        
        unnecessary_ports = self._check_unnecessary_ports(system)
        if unnecessary_ports:
            results.append(ComplianceResult(
                control_id="3.4.7",
                control_name=self.controls["3.4.7"]["name"],
                control_text=self.controls["3.4.7"]["text"],
                status="non_compliant",
                finding=f"Potentially unnecessary ports open: {unnecessary_ports}",
                remediation="Review and close unnecessary ports, disable unused services",
                severity="medium",
                evidence=[f"Open ports: {system.open_ports}"]
            ))
        
        external_services = self._check_external_services(system)
        if external_services:
            results.append(ComplianceResult(
                control_id="3.13.1",
                control_name=self.controls["3.13.1"]["name"],
                control_text=self.controls["3.13.1"]["text"],
                status="non_compliant",
                finding=f"External-facing services detected: {external_services}",
                remediation="Implement firewall rules, access controls, and monitoring for external-facing services",
                severity="high",
                evidence=[f"External services: {external_services}"]
            ))
        
        return results
    
    def _check_weak_services(self, system: SystemInfo) -> List[str]:
        """Check for potentially weak or insecure services"""
        weak_services = []
        risky_ports = {21: "FTP", 23: "Telnet", 135: "RPC", 139: "NetBIOS", 445: "SMB"}
        
        for port in system.open_ports:
            if port in risky_ports:
                service_name = system.services.get(port, risky_ports[port])
                weak_services.append(f"{service_name} (port {port})")
                
        return weak_services
    
    def _check_unnecessary_ports(self, system: SystemInfo) -> List[int]:
        """Check for potentially unnecessary open ports"""
        essential_ports = {22, 80, 443}
        return [port for port in system.open_ports if port not in essential_ports]
    
    def _check_external_services(self, system: SystemInfo) -> List[str]:
        """Check for services that might be externally accessible"""
        external_services = []
        external_ports = {21, 22, 23, 80, 443, 993, 995}
        
        for port in system.open_ports:
            if port in external_ports:
                service_name = system.services.get(port, f"Port {port}")
                external_services.append(service_name)
                
        return external_services

class SPRSCalculator:
    """Calculate SPRS (Supplier Performance Risk System) score"""
    
    def calculate_sprs_score(self, results: List[ComplianceResult]) -> Dict:
        """Calculate SPRS score based on compliance results"""
        logger.info("[SPRS] Calculating SPRS compliance score...")
        
        total_controls = len(self.get_all_control_ids())
        
        compliant = len([r for r in results if r.status == 'compliant'])
        non_compliant = len([r for r in results if r.status == 'non_compliant'])
        not_applicable = len([r for r in results if r.status == 'not_applicable'])
        not_assessed = len([r for r in results if r.status == 'not_assessed'])
        
        applicable_controls = total_controls - not_applicable
        if applicable_controls > 0:
            compliance_percentage = (compliant / applicable_controls) * 100
        else:
            compliance_percentage = 100
            
        base_score = 110
        
        high_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'high']) * 15
        medium_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'medium']) * 10
        low_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'low']) * 5
        
        total_deduction = high_severity_deduction + medium_severity_deduction + low_severity_deduction
        sprs_score = max(0, base_score - total_deduction)
        
        return {
            'sprs_score': sprs_score,
            'max_score': base_score,
            'compliance_percentage': round(compliance_percentage, 2),
            'total_controls': total_controls,
            'compliant': compliant,
            'non_compliant': non_compliant,
            'not_applicable': not_applicable,
            'not_assessed': not_assessed,
            'high_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'high']),
            'medium_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'medium']),
            'low_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'low'])
        }
    
    def get_all_control_ids(self) -> List[str]:
        """Get all NIST SP 800-171 control IDs"""
        return list(NIST80171Controls.CONTROLS.keys())

class ReportGenerator:
    """Generate compliance reports in multiple formats"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def generate_all_reports(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, topology: NetworkTopology = None):
        """Generate reports in all formats"""
        logger.info("[REPORTS] Generating compliance reports in multiple formats...")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        self.generate_html_report(systems, results, sprs_data, f"nistify_compliance_report_{timestamp}.html", topology)
        self.generate_json_report(systems, results, sprs_data, f"nistify_compliance_report_{timestamp}.json", topology)
        self.generate_text_report(systems, results, sprs_data, f"nistify_compliance_report_{timestamp}.txt")
        self.generate_poam_xlsx(results, f"nistify_poam_{timestamp}.xlsx")
        
        logger.info(f"[REPORTS] Reports generated in {self.output_dir}")
    
    def generate_html_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str, topology: NetworkTopology = None):
        """Generate HTML compliance report"""
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>NISTify 800-171 R2 - Compliance Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f8f9fa; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; text-align: center; }}
        .summary {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; margin: 20px 0; border-radius: 10px; }}
        .sprs-score {{ font-size: 36px; font-weight: bold; text-align: center; background: rgba(255,255,255,0.2); padding: 20px; border-radius: 10px; margin: 15px 0; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; background: white; border-radius: 10px; overflow: hidden; }}
        th, td {{ border: none; padding: 12px 15px; text-align: left; }}
        th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        .compliant {{ background-color: #d4edda !important; }}
        .non-compliant {{ background-color: #f8d7da !important; }}
        .high-severity {{ color: #dc3545; font-weight: bold; }}
        .medium-severity {{ color: #fd7e14; font-weight: bold; }}
        .low-severity {{ color: #28a745; font-weight: bold; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>NISTify 800-171 R2</h1>
        <p>NIST SP 800-171 Rev 2 Compliance Assessment Report</p>
        <p>Generated on: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    </div>
    
    <div class="summary">
        <h2>Executive Summary</h2>
        <div class="sprs-score">
            SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}
            <div style="font-size: 18px; margin-top: 10px;">
                Compliance: {sprs_data['compliance_percentage']}%
            </div>
        </div>
        <p>Systems Assessed: {len(systems)}</p>
        <p>Total Findings: {len(results)}</p>
        <p>High Severity: {sprs_data['high_severity_findings']}, Medium: {sprs_data['medium_severity_findings']}, Low: {sprs_data['low_severity_findings']}</p>
    </div>
    
    <h2>Scanned Systems</h2>
    <table>
        <tr>
            <th>Hostname</th>
            <th>IP Address</th>
            <th>OS Type</th>
            <th>Open Ports</th>
            <th>Last Scanned</th>
        </tr>
"""
        
        for system in systems:
            html_content += f"""
        <tr>
            <td><strong>{system.hostname}</strong></td>
            <td>{system.ip_address}</td>
            <td>{system.os_type}</td>
            <td>{', '.join(map(str, system.open_ports))}</td>
            <td>{system.last_scanned}</td>
        </tr>
"""
        
        html_content += """
    </table>
    
    <h2>Compliance Findings</h2>
    <table>
        <tr>
            <th>Control ID</th>
            <th>Control Name</th>
            <th>Status</th>
            <th>Severity</th>
            <th>Finding</th>
            <th>Remediation</th>
        </tr>
"""
        
        for result in results:
            status_class = result.status.replace('_', '-')
            severity_class = f"{result.severity}-severity"
            html_content += f"""
        <tr class="{status_class}">
            <td><strong>{result.control_id}</strong></td>
            <td>{result.control_name}</td>
            <td>{result.status.replace('_', ' ').title()}</td>
            <td class="{severity_class}">{result.severity.title()}</td>
            <td>{result.finding}</td>
            <td>{result.remediation}</td>
        </tr>
"""
        
        html_content += """
    </table>
    
    <div style="text-align: center; margin-top: 40px; padding: 20px; background: #f8f9fa; border-radius: 10px;">
        <p style="color: #666; margin: 0;">
            Generated by NISTify 800-171 R2 v1.2.0 | 
            Comprehensive NIST SP 800-171 Rev 2 Compliance Assessment Tool
        </p>
    </div>
</body>
</html>
        """
        
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def generate_json_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str, topology: NetworkTopology = None):
        """Generate JSON compliance report"""
        report_data = {
            "metadata": {
                "generated_on": datetime.datetime.now().isoformat(),
                "standard": "NIST SP 800-171 Rev 2",
                "tool": "NISTify 800-171 R2",
                "version": "1.2.0"
            },
            "sprs_score": sprs_data,
            "scanned_systems": [asdict(system) for system in systems],
            "compliance_results": [asdict(result) for result in results]
        }
        
        if topology:
            report_data["network_topology"] = {
                "nodes": topology.nodes,
                "edges": topology.edges,
                "subnets": topology.subnets,
                "gateways": topology.gateways,
                "diagram_path": topology.network_diagram_path
            }
        
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
    
    def generate_text_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str):
        """Generate text compliance report"""
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write("╔══════════════════════════════════════════════════════════════════════════════════════════════╗\n")
            f.write("║                          NISTify 800-171 R2 COMPLIANCE REPORT                               ║\n")
            f.write("╚══════════════════════════════════════════════════════════════════════════════════════════════╝\n\n")
            
            f.write(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Standard: NIST SP 800-171 Rev 2\n")
            f.write(f"Tool: NISTify 800-171 R2 v1.2.0\n\n")
            
            f.write("EXECUTIVE SUMMARY\n")
            f.write("=" * 50 + "\n")
            f.write(f"SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}\n")
            f.write(f"Compliance Percentage: {sprs_data['compliance_percentage']}%\n")
            f.write(f"Systems Assessed: {len(systems)}\n")
            f.write(f"Total Findings: {len(results)}\n")
            f.write(f"High Severity: {sprs_data['high_severity_findings']}\n")
            f.write(f"Medium Severity: {sprs_data['medium_severity_findings']}\n")
            f.write(f"Low Severity: {sprs_data['low_severity_findings']}\n\n")
            
            f.write("SCANNED SYSTEMS\n")
            f.write("=" * 30 + "\n")
            for system in systems:
                f.write(f"Hostname: {system.hostname}\n")
                f.write(f"IP Address: {system.ip_address}\n")
                f.write(f"OS Type: {system.os_type}\n")
                f.write(f"Open Ports: {', '.join(map(str, system.open_ports))}\n")
                f.write(f"Last Scanned: {system.last_scanned}\n\n")
            
            f.write("COMPLIANCE FINDINGS\n")
            f.write("=" * 35 + "\n")
            for result in results:
                f.write(f"Control ID: {result.control_id}\n")
                f.write(f"Control Name: {result.control_name}\n")
                f.write(f"Status: {result.status.replace('_', ' ').title()}\n")
                f.write(f"Severity: {result.severity.title()}\n")
                f.write(f"Finding: {result.finding}\n")
                f.write(f"Remediation: {result.remediation}\n")
                if result.evidence:
                    f.write(f"Evidence: {'; '.join(result.evidence)}\n")
                f.write("\n" + "-" * 80 + "\n\n")
            
            f.write("\n" + "="*80 + "\n")
            f.write("Generated by NISTify 800-171 R2 v1.2.0\n")
            f.write("Comprehensive NIST SP 800-171 Rev 2 Compliance Assessment Tool\n")
            f.write("="*80 + "\n")
    
    def generate_poam_xlsx(self, results: List[ComplianceResult], filename: str):
        """Generate Plan of Action and Milestones (POA&M) Excel document"""
        logger.info("[POAM] Generating POA&M Excel document...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "NISTify POA&M"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        low_fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
        
        headers = [
            "Control Number", "Control Name", "Control Text", "Status", "Severity",
            "Deficiency Identified", "Remediation Steps", "Target Date", 
            "Responsible Party", "Status Notes", "Evidence"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        for result in results:
            if result.status == 'non_compliant':
                ws.cell(row, 1, result.control_id)
                ws.cell(row, 2, result.control_name)
                ws.cell(row, 3, result.control_text)
                ws.cell(row, 4, result.status.replace('_', ' ').title())
                
                severity_cell = ws.cell(row, 5, result.severity.title())
                if result.severity == 'high':
                    severity_cell.fill = high_fill
                elif result.severity == 'medium':
                    severity_cell.fill = medium_fill
                else:
                    severity_cell.fill = low_fill
                
                ws.cell(row, 6, result.finding)
                ws.cell(row, 7, result.remediation)
                
                target_date = datetime.datetime.now()
                if result.severity == 'high':
                    target_date += datetime.timedelta(days=30)
                elif result.severity == 'medium':
                    target_date += datetime.timedelta(days=90)
                else:
                    target_date += datetime.timedelta(days=180)
                
                ws.cell(row, 8, target_date.strftime("%Y-%m-%d"))
                ws.cell(row, 9, "IT Security Team")
                ws.cell(row, 10, "Open")
                ws.cell(row, 11, '; '.join(result.evidence) if result.evidence else "")
                
                row += 1
        
        column_widths = [15, 30, 50, 15, 10, 40, 50, 12, 20, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = width
        
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"
        
        wb.save(self.output_dir / filename)
        logger.info(f"[POAM] POA&M Excel document generated: {filename}")

class ComplianceScanner:
    """Main NISTify compliance scanner orchestrator"""
    
    def __init__(self, output_dir: str = "reports"):
        self.scanner = NetworkScanner()
        self.assessor = ComplianceAssessor()
        self.sprs_calculator = SPRSCalculator()
        self.report_generator = ReportGenerator(output_dir)
        
    def scan_and_assess(self, network_ranges: List[str], ports: str = None, generate_topology: bool = True) -> Tuple[List[SystemInfo], List[ComplianceResult], Dict, NetworkTopology]:
        """Perform complete scan and assessment"""
        logger.info("[START] Starting NISTify 800-171 R2 compliance assessment...")
        
        all_systems = []
        all_results = []
        topology = None
        
        for network_range in network_ranges:
            logger.info(f"[SCAN] Scanning network range: {network_range}")
            systems = self.scanner.scan_network(network_range, ports)
            all_systems.extend(systems)
        
        if generate_topology and all_systems:
            topology = self.scanner.discover_topology(all_systems)
            diagram_path = str(Path(self.report_generator.output_dir) / "nistify_network_topology.png")
            self.scanner.create_network_diagram(topology, diagram_path)
        
        logger.info(f"[ASSESS] Assessing {len(all_systems)} systems for NIST SP 800-171 compliance")
        for system in all_systems:
            results = self.assessor.assess_system(system)
            all_results.extend(results)
        
        sprs_data = self.sprs_calculator.calculate_sprs_score(all_results)
        
        return all_systems, all_results, sprs_data, topology
    
    def generate_reports(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, topology: NetworkTopology = None):
        """Generate all compliance reports"""
        self.report_generator.generate_all_reports(systems, results, sprs_data, topology)

def main():
    """Main entry point"""
    print_banner()
    
    parser = argparse.ArgumentParser(
        description="NISTify 800-171 R2 - NIST SP 800-171 Rev 2 Compliance Scanner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python nistify800-171r2.py 192.168.1.0/24
  python nistify800-171r2.py 10.0.0.0/8 192.168.0.0/16 --verbose
  python nistify800-171r2.py 172.16.0.0/12 --ports "22,80,443,3389" --no-topology
        """
    )
    
    parser.add_argument("networks", nargs="+", 
                       help="Network ranges to scan (e.g., 192.168.1.0/24)")
    parser.add_argument("--ports", default="22,23,53,80,135,139,443,445,993,995", 
                       help="Comma-separated ports to scan (default: common ports)")
    parser.add_argument("--output-dir", default="nistify_reports", 
                       help="Output directory for reports (default: nistify_reports)")
    parser.add_argument("--no-topology", action="store_true", 
                       help="Skip network topology generation for faster scanning")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.info("[CONFIG] Verbose logging enabled")
    
    valid_networks = []
    logger.info("[CONFIG] Validating network ranges...")
    for network in args.networks:
        try:
            ipaddress.ip_network(network, strict=False)
            valid_networks.append(network)
            logger.info(f"[CONFIG] Valid network range: {network}")
        except ValueError:
            logger.error(f"[ERROR] Invalid network range: {network}")
            continue
    
    if not valid_networks:
        logger.error("[ERROR] No valid network ranges provided")
        sys.exit(1)
    
    scanner = ComplianceScanner(args.output_dir)
    
    try:
        logger.info("[START] Starting NISTify 800-171 R2 compliance assessment")
        start_time = datetime.datetime.now()
        
        systems, results, sprs_data, topology = scanner.scan_and_assess(
            valid_networks, args.ports, not args.no_topology
        )
        
        end_time = datetime.datetime.now()
        duration = end_time - start_time
        
        logger.info(f"[COMPLETE] Assessment complete in {duration.total_seconds():.1f} seconds")
        logger.info(f"[RESULTS] Found {len(systems)} systems with {len(results)} findings")
        logger.info(f"[SPRS] SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}")
        
        if topology:
            logger.info(f"[TOPOLOGY] Network topology: {len(topology.nodes)} nodes, {len(topology.edges)} connections")
            logger.info(f"[TOPOLOGY] Discovered subnets: {', '.join(topology.subnets)}")
        
        scanner.generate_reports(systems, results, sprs_data, topology)
        
        print(f"\n{'='*90}")
        print("NISTify 800-171 R2 COMPLIANCE ASSESSMENT COMPLETE")
        print(f"{'='*90}")
        print(f"Assessment Duration: {duration.total_seconds():.1f} seconds")
        print(f"Systems Scanned: {len(systems)}")
        print(f"Compliance Findings: {len(results)}")
        print(f"SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}")
        print(f"Compliance Rate: {sprs_data['compliance_percentage']}%")
        print(f"High Severity Issues: {sprs_data['high_severity_findings']}")
        print(f"Medium Severity Issues: {sprs_data['medium_severity_findings']}")
        print(f"Low Severity Issues: {sprs_data['low_severity_findings']}")
        
        if topology:
            print(f"\nNetwork Topology Analysis:")
            print(f"   Total Nodes: {len(topology.nodes)}")
            print(f"   Network Connections: {len(topology.edges)}")
            print(f"   Subnets Discovered: {len(topology.subnets)}")
            print(f"   Gateways Identified: {len(topology.gateways)}")
        
        print(f"\nReports generated in: {args.output_dir}")
        print("   HTML Report: nistify_compliance_report_*.html")
        print("   JSON Report: nistify_compliance_report_*.json")
        print("   Text Report: nistify_compliance_report_*.txt")
        print("   POA&M Document: nistify_poam_*.xlsx")
        
        if topology and topology.network_diagram_path:
            print("   Network Topology Diagram: nistify_network_topology.png")
        
        print(f"\nThank you for using NISTify 800-171 R2!")
        print("   For support and updates: https://github.com/yourusername/nistify800-171r2")
        
    except KeyboardInterrupt:
        logger.info("\n[INTERRUPTED] Assessment interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"[ERROR] Assessment failed: {e}")
        if args.verbose:
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()
